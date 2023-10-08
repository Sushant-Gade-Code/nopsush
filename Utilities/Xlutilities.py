
import openpyxl
class Xlutility:

    @staticmethod
    def getRowsCount(file,sheetname):
        workBook=openpyxl.load_workbook(file)
        sheet=workBook[sheetname]
        rows=sheet.max_row
        return rows

    @staticmethod
    def getColumnCount(file, sheetname):
        workBook = openpyxl.load_workbook(file)
        sheet = workBook[sheetname]
        columns = sheet.max_column
        return columns

    @staticmethod
    def getReadData(file, sheetname,rownum,column):
        workBook = openpyxl.load_workbook(file)
        sheet = workBook[sheetname]
        return sheet.cell(rownum,column).value

    @staticmethod
    def getWritedata(file, sheetname, rownum, column,data):
        workBook = openpyxl.load_workbook(file)
        sheet = workBook[sheetname]
        (sheet.cell(rownum,column).value)=data
        workBook.save(file)





